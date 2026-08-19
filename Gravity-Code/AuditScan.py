#!/usr/bin/env python3
"""
AuditScan - Ultra-fast Discovery Accelerator Risky Keyword Scanner
------------------------------------------------------------------
Scans Discovery Accelerator search logs (.csv, .xlsx, .txt) for risky keywords, calculates threat scores, and generates a color-coded Excel audit report. """

import os
import sys
import re
import csv

try:
    from tkinter import Tk, filedialog
except ImportError:
    Tk = filedialog = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    Workbook = load_workbook = None

# ==============================================================================
# THREAT PATTERNS & KEYWORDS
# ==============================================================================

KEYWORDS = {
    "exfiltration": ("export", "download all", "extract", "dump", "copy all", "remove audit", "unmonitored", "disable logging", "pst export"),
    "privilege": ("admin", "root", "privileged", "elevated", "bypass", "override", "superuser"),
    "confidential": ("salary", "ssn", "social security", "compensation", "bonus", "confidential", "secret", "medical", "password", "passport","Drivers license", "Date of birth","ITIN","Mother's maiden name", "Retirement", "Pension", "Banking","Credit Card", "PIN", "IBAN", "CVV"),
    "circumvention": ("evade", "avoid logs", "delete logs", "cover tracks", "hide search", "disable audit"),
    "unauthorized": ("executive email", "board mailbox", "legal archive", "ceo", "cfo", "whistleblower", "investigation"),
    "custom": ("regulatory", "compliance", "personal email", "legal record", "suspicious")
}

PATTERNS = {
    "bulk_access": re.compile(r"(all\s+emails|entire\s+archive|full\s+mailbox)", re.I),
    "time_anomaly": re.compile(r"(midnight|3am|after\s+hours|off[- ]hours)", re.I),
    "intent_anomaly": re.compile(r"(no\s+one\s+should\s+know|secret|hidden|do\s+not\s+tell)", re.I)
}

def load_queries(path: str) -> list:
    """Extract search queries from .csv, .xlsx, or .txt files."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            r = csv.reader(f)
            header = [h.strip().lower() for h in next(r, [])]
            col = header.index("query") if "query" in header else 0
            return [{"query": row[col].strip(), "row": i} for i, row in enumerate(r, 2) if len(row) > col and row[col].strip()]
    elif ext in (".xlsx", ".xls"):
        if not load_workbook:
            raise RuntimeError("openpyxl is required for Excel files: pip install openpyxl")
        ws = load_workbook(path, data_only=True).active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(h).strip().lower() if h else "" for h in rows[0]]
        col = header.index("query") if "query" in header else 0
        return [{"query": str(row[col]).strip(), "row": i} for i, row in enumerate(rows[1:], 2) if len(row) > col and row[col] and str(row[col]).strip()]
    else:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return [{"query": line.strip(), "row": i} for i, line in enumerate(f, 1) if line.strip() and not line.strip().startswith("#")]


def score_query(query: str) -> dict:
    """Score search query against threat patterns."""
    q_low = query.lower()
    hits = [f"{w} ({cat})" for cat, words in KEYWORDS.items() for w in words if w in q_low]
    hits += [f"pattern:{name}" for name, pat in PATTERNS.items() if pat.search(q_low)]

    score = sum(2 if not h.startswith("pattern:") else 3 for h in hits) + (1 if len(query) > 200 else 0)
    risk = "HIGH" if score >= 7 else ("MEDIUM" if score >= 4 else "LOW")
    return {"query": query, "score": score, "risk": risk, "hits": hits}


def export_report(results: list, path: str):
    """Generate color-coded Excel audit report."""
    if not Workbook:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Report"

    styles = {
        "HIGH": (PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), Font(color="FFFFFF", bold=True)),
        "MEDIUM": (PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"), Font(color="000000", bold=True)),
        "LOW": (PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), Font(color="000000"))
    }
    head_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

    ws.append(["#", "Search Query", "Risk Level", "Score", "Detected Keywords & Patterns"])
    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = head_fill, head_font, Alignment(horizontal="center")

    for i, r in enumerate(results, 1):
        ws.append([i, r["query"], r["risk"], r["score"], ", ".join(r["hits"]) or "(none)"])
        row = ws[i + 1]
        for c in row:
            c.border = border
        row[2].fill, row[2].font = styles.get(r["risk"], (None, None))
        row[2].alignment = row[3].alignment = Alignment(horizontal="center")

    for col, width in zip("ABCDE", [8, 55, 14, 10, 50]):
        ws.column_dimensions[col].width = width

    wb.save(path)


def main():
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    print("=" * 60 + "\n  AUDITSCAN - DISCOVERY ACCELERATOR LOG SCANNER\n" + "=" * 60)

    # File selection
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    elif Tk:
        root = Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename(title="Select Log File", filetypes=[("Log Files", "*.csv *.xlsx *.xls *.txt *.log"), ("All", "*.*")])
        root.destroy()
    else:
        print("Usage: python AuditScan.py <logfile>")
        return

    if not filepath or not os.path.exists(filepath):
        print("No valid file selected. Exiting.")
        return

    print(f"[*] Scanning: {filepath}")
    records = load_queries(filepath)
    if not records:
        print("No queries found.")
        return

    results = [dict(score_query(r["query"]), row=r["row"]) for r in records]
    risky = [r for r in results if r["risk"] in ("HIGH", "MEDIUM")]

    if risky:
        print("\n" + "=" * 60 + f"\n  RISKY QUERIES DETECTED ({len(risky)} Flagged)\n" + "=" * 60)
        for i, r in enumerate(risky, 1):
            badge = "[HIGH]" if r["risk"] == "HIGH" else "[MED]"
            print(f"{i:3d}. {badge} Score: {r['score']} (Row {r['row']}): \"{r['query']}\"\n     Hits: {', '.join(r['hits'])}")

    total = len(results)
    high = sum(1 for r in results if r["risk"] == "HIGH")
    med = sum(1 for r in results if r["risk"] == "MEDIUM")
    low = total - high - med

    print("\n" + "=" * 60 + f"\n  SUMMARY: {total} Queries Scanned")
    print(f"  High Risk (RED):     {high:4d} ({(high/total*100):.1f}%)")
    print(f"  Medium Risk (AMBER): {med:4d} ({(med/total*100):.1f}%)")
    print(f"  Low Risk (GREEN):    {low:4d} ({(low/total*100):.1f}%)\n" + "=" * 60)

    out_excel = os.path.splitext(filepath)[0] + "_audit_report.xlsx"
    export_report(results, out_excel)
    print(f"[+] Report saved: {out_excel}")

    try:
        os.startfile(out_excel) if sys.platform == "win32" else os.system(f'open "{out_excel}"')
    except Exception:
        pass


if __name__ == "__main__":
    main()
