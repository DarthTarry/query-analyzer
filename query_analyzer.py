# ===========================================
# QUERY ANALYZER - DOCUMENTATION
# ===========================================

"""
QUERY ANALYZER - Security Threat Detection System

OVERVIEW:
Query Analyzer is a security threat detection system that analyzes text queries for suspicious activity indicators. It scores queries based on risk patterns, keyword matches, and behavioral anomalies, then generates comprehensive reports 
with visual risk indicators.

KEY FEATURES:
✅ Keyword-based threat detection (5 categories)
✅ Regex pattern matching for complex threats
✅ Risk scoring algorithm (0-12+ points)
✅ File-based batch processing
✅ Color-coded Excel reports (RED/AMBER/WHITE)
✅ Console output with alerts
✅ Automatic file opening
✅ Cross-platform support (Windows/Mac/Linux)

---

SCORING ALGORITHM:

Detection Type     | Points | Description
────────────────────────────────────────────────────────────
Keyword Match      | +2     | Each suspicious keyword found
Regex Pattern      | +3     | Complex threat pattern detected
Length Anomaly     | +1     | Query exceeds 200 characters
Time Anomaly       | +1     | Executed outside 6 AM - 10 PM

RISK CLASSIFICATION:
Score ≥ 7:    HIGH RISK    (🔴 RED)    - Immediate investigation needed
Score 4-6:    MEDIUM RISK  (🟠 AMBER)  - Review and monitor
Score < 4:    LOW RISK     (⚪ WHITE)  - Normal business query

---

THREAT KEYWORD CATEGORIES:

1. DATA EXFILTRATION (Unauthorized Data Access):
   Keywords: export, download all, extract, dump, copy all, remove audit, 
             unmonitored, disable logging
   Detects: Attempts to access or exfiltrate data en masse
   
2. ACCESS PRIVILEGE (Unauthorized Elevation):
   Keywords: admin, root, privileged, elevated, bypass
   Detects: Attempts to bypass security or escalate privileges
   
3. CONFIDENTIAL INFORMATION (Sensitive Data Targeting):
   Keywords: salary, ssn, social security, medical, private
   Detects: Queries targeting known sensitive data types
   
4. CIRCUMVENTION (Security Bypass):
   Keywords: evade, avoid logs, delete logs, cover tracks
   Detects: Attempts to hide activity or bypass security monitoring
   
5. UNAUTHORIZED TARGETS (High-Value Targets):
   Keywords: executive email, board mailbox, legal archive
   Detects: Queries targeting executive or legal resources

---

REGEX PATTERN DETECTION:

Pattern 1: Bulk Access
  Regex: (all\\s+emails|entire\\s+archive|full\\s+mailbox)
  Triggers: Attempts to access complete data sets
  Score: +3 points
  
Pattern 2: Time Anomaly
  Regex: (midnight|3am|after hours)
  Triggers: Suspicious activity time indicators
  Score: +3 points
  
Pattern 3: Intent Anomaly
  Regex: (no one should know|secret|hidden)
  Triggers: Explicit intent to conceal activity
  Score: +3 points

---

HOW IT WORKS:

Step 1: File Input
  - User selects .txt file via dialog
  - File contains one query per line
  
Step 2: Query Analysis
  For each query, the script:
  1. Scans keywords against 5 threat categories
  2. Applies regex patterns for complex threats
  3. Checks anomalies (length, time)
  4. Calculates total score
  5. Classifies risk level (HIGH/MEDIUM/LOW)
  6. Records indicators (which patterns matched)
  
Step 3: Report Generation
  Creates Excel file with:
  - Formatted headers (blue background, white text)
  - Color-coded risk cells (RED/AMBER/WHITE)
  - Detailed indicators for each query
  - Sortable/filterable data
  
Step 4: Output
  ✅ Console output with alerts
  ✅ Excel file (.xlsx) automatically opened
  ✅ All results saved for audit trail

---

USAGE INSTRUCTIONS:

Running the Script:
  Method 1 (Python):
    python c:\\Users\\arunv\\query_analyzer.py
    
  Method 2 (Executable):
    Double-click: C:\\Users\\arunv\\Desktop\\query_analyzer.exe

Step-by-Step Usage:
  1. Launch the script
  2. Select your queries file (queries.txt)
  3. Review console output with immediate alerts
  4. Check the Excel report that auto-opens
  5. Export results for audit/compliance

Input File Format:
  Create a .txt file with one query per line:
  
  export all customer data
  retrieve employee salary database
  admin access to board mailbox
  normal daily report

---

EXAMPLE SCORING:

Example 1: "export all emails"
  - Keyword: "export" (data_exfiltration) = +2
  - Pattern: "all emails" (bulk_access) = +3
  - Total Score: 5 → MEDIUM RISK ⚠️

Example 2: "admin password delete logs midnight"
  - Keyword: "admin" (access_privilege) = +2
  - Keyword: "delete logs" (circumvention) = +2
  - Pattern: "midnight" (time_anomaly) = +3
  - Time anomaly (after 10 PM) = +1
  - Total Score: 8 → HIGH RISK ⚠️

Example 3: "generate quarterly sales report for Q3"
  - No keywords matched = +0
  - No patterns detected = +0
  - Query length normal = +0
  - Executed during business hours = +0
  - Total Score: 0 → LOW RISK ✓

---

OUTPUT FILES:

Excel Report Format:
  File naming: input_base_analysis_report.xlsx
  
  Columns:
  - Query #: Sequential query number
  - Query: The actual query text
  - Risk Level: HIGH/MEDIUM/LOW with color coding
  - Score: Numeric risk score
  - Indicators: List of detected threat patterns

Console Output Example:
  === Query 1 ===
  Query: export all emails
  Risk Level: MEDIUM
  Score: 5
  Indicators:
   - Keyword match: export (data_exfiltration)
   - Pattern match: bulk_access

---

CUSTOMIZATION GUIDE:

Adding New Keywords:
  Edit SUSPICIOUS_KEYWORDS dictionary:
  
  SUSPICIOUS_KEYWORDS = {
      "data_exfiltration": [...],
      "your_new_category": ["keyword1", "keyword2"],  # ← Add here
  }

Adding New Regex Patterns:
  Edit SUSPICIOUS_REGEX dictionary:
  
  SUSPICIOUS_REGEX = {
      "bulk_access": r"(...)",
      "your_pattern": r"(pattern1|pattern2)",  # ← Add here
  }

Changing Risk Thresholds:
  Modify score_query() function:
  
  if score >= 7:      # Change threshold
      risk = "HIGH"
  elif score >= 4:    # Change threshold
      risk = "MEDIUM"

Changing Colors:
  Edit PatternFill definitions:
  
  high_risk_fill = PatternFill(start_color="FF0000", ...)  # Red
  medium_risk_fill = PatternFill(start_color="FFC000", ...)  # Amber
  low_risk_fill = PatternFill(start_color="FFFFFF", ...)  # White

---

SECURITY CONSIDERATIONS:

This Tool Is Designed For:
  ✅ Security audit and compliance monitoring
  ✅ Threat detection and risk assessment
  ✅ Insider threat identification
  ✅ Query auditing and logging

Limitations:
  ⚠️ Keyword-based detection (can be evaded with synonyms)
  ⚠️ Regex patterns may have false positives
  ⚠️ Does not analyze actual data access
  ⚠️ Should be used alongside other security controls

Recommendations:
  - Use as part of a multi-layered security strategy
  - Combine with actual system access logs
  - Review HIGH risk queries with security team
  - Regularly update keyword/pattern databases
  - Test false positive rates in your environment

---

TROUBLESHOOTING:

Issue: "File not found"
  Cause: Incorrect file path or file doesn't exist
  Solution: Select file using the dialog box

Issue: Excel won't open automatically
  Cause: openpyxl not installed
  Solution: pip install openpyxl

Issue: No keywords detected
  Cause: Keywords may be in different case or form
  Solution: Keywords are case-insensitive; check exact spelling

Issue: All queries marked HIGH risk
  Cause: Thresholds may be too low
  Solution: Adjust score >= 7 threshold higher

---

TECHNICAL INFORMATION:

Python Version: 3.7+
Dependencies: openpyxl, tkinter (standard library)
Platforms: Windows, macOS, Linux

Key Functions:
  - score_query(query: str) -> dict
    Analyzes single query, returns {query, score, risk_level, indicators}
    
  - analyze_queries_from_file(filename: str) -> None
    Reads queries from file, creates Excel workbook, applies styling,
    generates console output, auto-opens Excel file

---

VERSION HISTORY:
- v1.0: Initial release with 5 keyword categories, 3 regex patterns,
  color-coded Excel output, cross-platform support

STATUS: Production Ready ✅
Last Updated: 2026-08-15

For more information, see: QUERY_ANALYZER_DOCUMENTATION.md
"""

import re
import sys
import os
import csv
import subprocess
from datetime import datetime
from tkinter import Tk, Toplevel, Label, Button, Frame, Text, filedialog, messagebox, simpledialog

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:  # pragma: no cover - handled gracefully for environments without the dependency
    Workbook = None
    load_workbook = None
    PatternFill = Font = Alignment = Border = Side = None


def ensure_openpyxl():
    """Raise a clear error if Excel support is unavailable."""
    if Workbook is None or load_workbook is None or PatternFill is None:
        raise RuntimeError(
            "openpyxl is required for Excel import/export features. "
            "Install it with: python -m pip install openpyxl"
        )


# Provide a helpful runtime error only when workbook functionality is used.
def _openpyxl_unavailable(*args, **kwargs):
    ensure_openpyxl()

# -----------------------------------------
# Suspicious pattern definitions
# -----------------------------------------

SUSPICIOUS_KEYWORDS = {
    "data_exfiltration": ["export", "download all", "extract", "dump", "copy all", "remove audit","unmonitored","disable logging"],
    "access_privilege": ["admin", "root", "privileged", "elevated", "bypass"],
    "confidential_info": ["salary", "ssn", "social security", "medical", "private"],
    "circumvention": ["evade", "avoid logs", "delete logs", "cover tracks"],
    "unauthorized_targets": ["executive email", "board mailbox", "legal archive"],
}

SUSPICIOUS_REGEX = {
    "bulk_access": r"(all\s+emails|entire\s+archive|full\s+mailbox)",
    "time_anomaly": r"(midnight|3am|after hours)",
    "intent_anomaly": r"(no one should know|secret|hidden)",
}

CUSTOM_KEYWORDS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "custom_keywords.txt")


def load_saved_custom_keywords() -> list:
    """Load custom keywords from the local save file so they persist between runs."""
    if not os.path.exists(CUSTOM_KEYWORDS_FILE):
        return []

    try:
        with open(CUSTOM_KEYWORDS_FILE, "r", encoding="utf-8") as file:
            loaded = [line.strip().lower() for line in file if line.strip()]
        return list(dict.fromkeys(loaded))
    except Exception:
        return []


def save_custom_keywords(keywords: list) -> None:
    """Persist custom keywords for future runs."""
    cleaned = []
    for keyword in keywords:
        normalized = keyword.strip().lower()
        if normalized and normalized not in cleaned:
            cleaned.append(normalized)

    with open(CUSTOM_KEYWORDS_FILE, "w", encoding="utf-8") as file:
        for keyword in cleaned:
            file.write(f"{keyword}\n")


def prompt_for_keyword_updates(parent=None) -> list:
    """Prompt the user for custom suspicious keywords without creating a second Tk root."""
    keyword_summary = "\n\n".join(
        f"{category}: {', '.join(words)}"
        for category, words in SUSPICIOUS_KEYWORDS.items()
        if category != "custom_keywords"
    )

    info_message = (
        "Current suspicious keyword categories:\n\n"
        + keyword_summary
        + "\n\n"
        + "These default terms are used before any custom additions."
    )
    messagebox.showinfo("Suspicious Keywords", info_message)

    wants_to_add = messagebox.askyesno(
        "Add custom keywords?",
        "Would you like to add any custom suspicious phrases before running the analysis?",
    )

    if not wants_to_add:
        saved_keywords = load_saved_custom_keywords()
        if saved_keywords:
            SUSPICIOUS_KEYWORDS["custom_keywords"] = saved_keywords
        else:
            SUSPICIOUS_KEYWORDS.pop("custom_keywords", None)
        return saved_keywords

    if parent is None:
        parent = Tk()
        parent.withdraw()

    custom_keywords = []
    existing = load_saved_custom_keywords()
    if existing:
        custom_keywords.extend(existing)

    while True:
        prompt = "Enter a suspicious keyword or phrase. Leave blank to finish:"
        new_keyword = simpledialog.askstring(
            "Add Custom Keywords",
            prompt,
            parent=parent,
        )

        if new_keyword is None:
            break

        cleaned = new_keyword.strip().lower()
        if not cleaned:
            break

        if cleaned not in custom_keywords:
            custom_keywords.append(cleaned)

    if custom_keywords:
        save_custom_keywords(custom_keywords)
        SUSPICIOUS_KEYWORDS["custom_keywords"] = custom_keywords
    else:
        SUSPICIOUS_KEYWORDS.pop("custom_keywords", None)
        if os.path.exists(CUSTOM_KEYWORDS_FILE):
            try:
                os.remove(CUSTOM_KEYWORDS_FILE)
            except OSError:
                pass

    if parent and parent.winfo_exists():
        try:
            parent.mainloop()
        except Exception:
            pass

    return load_saved_custom_keywords()

# -----------------------------------------
# Scoring engine
# -----------------------------------------

def score_query(query: str) -> dict:
    query_lower = query.lower()
    score = 0
    hits = []

    # Keyword matches
    for category, words in SUSPICIOUS_KEYWORDS.items():
        for w in words:
            if w in query_lower:
                score += 2
                hits.append(f"Keyword match: {w} ({category})")

    # Regex matches
    for category, pattern in SUSPICIOUS_REGEX.items():
        if re.search(pattern, query_lower):
            score += 3
            hits.append(f"Pattern match: {category}")

    # Length anomaly (very long queries)
    if len(query) > 200:
        score += 1
        hits.append("Length anomaly: unusually long query")

    # Time anomaly (if used in real systems)
    current_hour = datetime.now().hour
    if current_hour < 6 or current_hour > 22:
        score += 1
        hits.append("Time anomaly: query executed outside normal hours")

    # Risk classification
    if score >= 7:
        risk = "HIGH"
    elif score >= 4:
        risk = "MEDIUM"
    else:
        risk = "LOW"

    return {
        "query": query,
        "score": score,
        "risk_level": risk,
        "indicators": hits
    }

# -----------------------------------------
# File input handler
# -----------------------------------------

def read_queries_from_excel(filename: str) -> list:
    """Read queries from Excel file (.xlsx or .xls).
    Looks for a 'query' column header and reads queries from that column."""
    ensure_openpyxl()
    try:
        wb = load_workbook(filename)
        ws = wb.active
        queries = []
        query_col_index = None
        
        # Find the 'query' column header
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for idx, header in enumerate(first_row):
            if header and isinstance(header, str) and header.lower() == "query":
                query_col_index = idx
                break
        
        if query_col_index is None:
            # Fallback to first column if 'query' header not found
            print("Warning: 'query' column header not found. Using first column.")
            query_col_index = 0
        
        # Read queries from the identified column
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) > query_col_index:
                cell_value = row[query_col_index]
                if cell_value and isinstance(cell_value, str):
                    query = cell_value.strip()
                    if query:  # Only add non-empty queries
                        queries.append(query)
        
        return queries
    except Exception as e:
        print(f"Error reading Excel file '{filename}': {e}")
        return []

def read_queries_from_text(filename: str) -> list:
    """Read queries from text file (.txt).
    Expects one query per line."""
    try:
        with open(filename, 'r') as f:
            queries = [line.strip() for line in f if line.strip()]
        return queries
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return []

def read_queries_from_csv(filename: str) -> list:
    """Read queries from CSV file (.csv).
    Looks for a 'query' column header and reads queries from that column."""
    try:
        queries = []
        query_col_index = None
        
        with open(filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            # Read header row
            header_row = next(reader, None)
            if header_row:
                # Find the 'query' column index
                for idx, header in enumerate(header_row):
                    if header.lower() == "query":
                        query_col_index = idx
                        break
            
            if query_col_index is None:
                # Fallback to first column if 'query' header not found
                print("Warning: 'query' column header not found. Using first column.")
                query_col_index = 0
            
            # Read queries from the identified column
            for row in reader:
                if row and len(row) > query_col_index:
                    query = row[query_col_index].strip()
                    if query:  # Only add non-empty queries
                        queries.append(query)
        
        return queries
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return []
    except Exception as e:
        print(f"Error reading CSV file '{filename}': {e}")
        return []

def calculate_kpi_metrics(results: list) -> dict:
    """Calculate KPI metrics from analysis results."""
    if not results:
        return {}
    
    total_queries = len(results)
    scores = [r['score'] for r in results]
    
    # Risk level counts
    high_count = sum(1 for r in results if r['risk_level'] == 'HIGH')
    medium_count = sum(1 for r in results if r['risk_level'] == 'MEDIUM')
    low_count = sum(1 for r in results if r['risk_level'] == 'LOW')
    
    # Score statistics
    avg_score = sum(scores) / total_queries if total_queries > 0 else 0
    min_score = min(scores) if scores else 0
    max_score = max(scores) if scores else 0
    
    # Indicator frequency analysis
    indicator_counts = {}
    for result in results:
        for indicator in result['indicators']:
            # Extract category from indicator string
            if "Keyword match:" in indicator:
                category = indicator.split('(')[-1].rstrip(')')
                key = f"Keyword: {category}"
            elif "Pattern match:" in indicator:
                key = f"Pattern: {indicator.split(': ')[1]}"
            else:
                key = indicator
            
            indicator_counts[key] = indicator_counts.get(key, 0) + 1
    
    # Sort indicators by frequency
    top_indicators = sorted(indicator_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    return {
        'total_queries': total_queries,
        'high_count': high_count,
        'medium_count': medium_count,
        'low_count': low_count,
        'high_percent': (high_count / total_queries * 100) if total_queries > 0 else 0,
        'medium_percent': (medium_count / total_queries * 100) if total_queries > 0 else 0,
        'low_percent': (low_count / total_queries * 100) if total_queries > 0 else 0,
        'avg_score': round(avg_score, 2),
        'min_score': min_score,
        'max_score': max_score,
        'score_range': max_score - min_score,
        'top_indicators': top_indicators
    }


def create_metrics_sheet(wb, metrics: dict):
    """Create a summary metrics sheet in the workbook."""
    ws = wb.create_sheet("KPI Summary", 0)  # Insert at beginning
    
    # Define styles
    title_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    
    metric_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    metric_font = Font(bold=True, size=11)
    
    high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    medium_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    white_font = Font(color="FFFFFF", bold=True, size=11)
    dark_font = Font(color="000000", bold=True, size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "SECURITY THREAT ANALYSIS - KPI SUMMARY"
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = center_alignment
    ws.row_dimensions[row].height = 25
    row += 2
    
    # Overall Statistics Section
    ws[f'A{row}'] = "OVERALL STATISTICS"
    ws[f'A{row}'].font = metric_font
    ws[f'A{row}'].fill = metric_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = "Total Queries Analyzed"
    ws[f'B{row}'] = metrics.get('total_queries', 0)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 1
    
    ws[f'A{row}'] = "Analysis Timestamp"
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 2
    
    # Risk Distribution Section
    ws[f'A{row}'] = "RISK DISTRIBUTION"
    ws[f'A{row}'].font = metric_font
    ws[f'A{row}'].fill = metric_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # High Risk
    ws[f'A{row}'] = "🔴 HIGH RISK"
    ws[f'B{row}'] = metrics.get('high_count', 0)
    ws[f'A{row}'].fill = high_fill
    ws[f'A{row}'].font = white_font
    ws[f'B{row}'].fill = high_fill
    ws[f'B{row}'].font = white_font
    ws[f'B{row}'].alignment = center_alignment
    row += 1
    
    # Medium Risk
    ws[f'A{row}'] = "🟠 MEDIUM RISK"
    ws[f'B{row}'] = metrics.get('medium_count', 0)
    ws[f'A{row}'].fill = medium_fill
    ws[f'A{row}'].font = dark_font
    ws[f'B{row}'].fill = medium_fill
    ws[f'B{row}'].font = dark_font
    ws[f'B{row}'].alignment = center_alignment
    row += 1
    
    # Low Risk
    ws[f'A{row}'] = "⚪ LOW RISK"
    ws[f'B{row}'] = metrics.get('low_count', 0)
    ws[f'A{row}'].fill = low_fill
    ws[f'A{row}'].font = dark_font
    ws[f'B{row}'].fill = low_fill
    ws[f'B{row}'].font = dark_font
    ws[f'B{row}'].alignment = center_alignment
    row += 2
    
    # Risk Percentages
    ws[f'A{row}'] = "RISK PERCENTAGES"
    ws[f'A{row}'].font = metric_font
    ws[f'A{row}'].fill = metric_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = "HIGH RISK %"
    ws[f'B{row}'] = round(metrics.get('high_percent', 0), 2)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 1
    
    ws[f'A{row}'] = "MEDIUM RISK %"
    ws[f'B{row}'] = round(metrics.get('medium_percent', 0), 2)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 1
    
    ws[f'A{row}'] = "LOW RISK %"
    ws[f'B{row}'] = round(metrics.get('low_percent', 0), 2)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 2
    
    # Scoring Statistics Section
    ws[f'A{row}'] = "SCORING STATISTICS"
    ws[f'A{row}'].font = metric_font
    ws[f'A{row}'].fill = metric_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = "Average Score"
    ws[f'B{row}'] = metrics.get('avg_score', 0)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 1
    
    ws[f'A{row}'] = "Minimum Score"
    ws[f'B{row}'] = metrics.get('min_score', 0)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 1
    
    ws[f'A{row}'] = "Maximum Score"
    ws[f'B{row}'] = metrics.get('max_score', 0)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 1
    
    ws[f'A{row}'] = "Score Range (Max - Min)"
    ws[f'B{row}'] = metrics.get('score_range', 0)
    ws[f'A{row}'].fill = metric_fill
    ws[f'A{row}'].font = metric_font
    row += 2
    
    # Top Threat Indicators Section
    ws[f'A{row}'] = "TOP THREAT INDICATORS"
    ws[f'A{row}'].font = metric_font
    ws[f'A{row}'].fill = metric_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    for indicator, count in metrics.get('top_indicators', []):
        ws[f'A{row}'] = indicator
        ws[f'B{row}'] = count
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = center_alignment
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # Apply borders to all cells
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row_cells:
            if cell.border.left.style is None:
                cell.border = border


def analyze_queries_from_file(filename: str):
    ensure_openpyxl()
    # Determine file type and read queries accordingly
    file_ext = os.path.splitext(filename)[1].lower()
    
    if file_ext in ['.xlsx', '.xls']:
        queries = read_queries_from_excel(filename)
    elif file_ext == '.csv':
        queries = read_queries_from_csv(filename)
    elif file_ext == '.txt':
        queries = read_queries_from_text(filename)
    else:
        print(f"Error: Unsupported file type '{file_ext}'. Please use .txt, .csv, .xlsx, or .xls files.")
        return
    
    if not queries:
        print(f"No queries found in '{filename}'.")
        return
    
    # Get the directory and create Excel file path
    file_dir = os.path.dirname(filename)
    file_base = os.path.splitext(os.path.basename(filename))[0]
    excel_filename = os.path.join(file_dir, f"{file_base}_analysis_report.xlsx")
    
    # Create a workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Query Analysis"
    
    # Define colors and styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    high_risk_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    medium_risk_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Amber
    low_risk_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    
    risk_font = Font(bold=True, size=11)
    red_font = Font(color="FFFFFF", bold=True, size=11)  # White text for red cells
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = ["Query #", "Query", "Risk Level", "Score", "Indicators"]
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50
    
    # Analyze all queries and store results
    results = []
    for i, query in enumerate(queries, 1):
        result = score_query(query)
        results.append(result)
        
        indicators_text = ", ".join(result["indicators"]) if result["indicators"] else "(none)"
        
        ws.append([
            i,
            result['query'],
            result['risk_level'],
            result['score'],
            indicators_text
        ])
        
        # Get the current row
        row_num = i + 1
        
        # Apply formatting to the row
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.border = border
            cell.alignment = alignment
            
            # Apply color coding to Risk Level column (column C)
            if col_num == 3:  # Risk Level column
                if result['risk_level'] == "HIGH":
                    cell.fill = high_risk_fill
                    cell.font = red_font
                    ws.row_dimensions[row_num].height = 25
                elif result['risk_level'] == "MEDIUM":
                    cell.fill = medium_risk_fill
                    cell.font = risk_font
                    ws.row_dimensions[row_num].height = 25
                else:
                    cell.fill = low_risk_fill
                cell.alignment = center_alignment
        
        # Print to console
        if result["risk_level"] == "HIGH":
            print("\n" + "!" * 60)
            print("WARNING: HIGH RISK QUERY DETECTED!")
            print("!" * 60)
        
        print(f"\n=== Query {i} ===")
        print(f"Query: {result['query']}")
        print(f"Risk Level: {result['risk_level']}")
        print(f"Score: {result['score']}")
        print("Indicators:")
        
        if result["indicators"]:
            for indicator in result["indicators"]:
                print(f" - {indicator}")
        else:
            print(" (none)")
        
        if result["risk_level"] == "HIGH":
            print("!" * 60)
    
    # Calculate and add KPI metrics
    metrics = calculate_kpi_metrics(results)
    create_metrics_sheet(wb, metrics)
    
    # Print KPI summary to console
    print("\n" + "=" * 60)
    print("KPI SUMMARY")
    print("=" * 60)
    print(f"Total Queries: {metrics['total_queries']}")
    print(f"High Risk: {metrics['high_count']} ({metrics['high_percent']:.1f}%)")
    print(f"Medium Risk: {metrics['medium_count']} ({metrics['medium_percent']:.1f}%)")
    print(f"Low Risk: {metrics['low_count']} ({metrics['low_percent']:.1f}%)")
    print(f"Average Score: {metrics['avg_score']}")
    print(f"Score Range: {metrics['min_score']} - {metrics['max_score']}")
    print("=" * 60)
    
    # Save the workbook
    wb.save(excel_filename)
    
    print(f"\nAnalysis complete! Results saved to: {excel_filename}")
    
    # Open the Excel file
    try:
        if sys.platform == 'win32':
            os.startfile(excel_filename)
        elif sys.platform == 'darwin':  # macOS
            subprocess.Popen(['open', excel_filename])
        else:  # Linux
            subprocess.Popen(['xdg-open', excel_filename])
        print("Opening Excel file...")
    except Exception as e:
        print(f"Could not open file automatically: {e}")

# -----------------------------------------
# Example usage
# -----------------------------------------

if __name__ == "__main__":
    root = Tk()
    root.withdraw()

    custom_keyword_updates = prompt_for_keyword_updates(root)
    if custom_keyword_updates:
        SUSPICIOUS_KEYWORDS["custom_keywords"] = list(dict.fromkeys(custom_keyword_updates))
    else:
        SUSPICIOUS_KEYWORDS.pop("custom_keywords", None)

    # Open file dialog
    filename = filedialog.askopenfilename(
        parent=root,
        title="Select a queries file",
        filetypes=[("All supported files", "*.xlsx *.xls *.csv *.txt"), ("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("Text files", "*.txt"), ("All files", "*.*")],
        defaultextension=".xlsx"
    )

    # Analyze if a file was selected
    if filename:
        analyze_queries_from_file(filename)
    else:
        print("No file selected. Exiting.")

    root.destroy()
